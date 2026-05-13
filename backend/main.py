from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import httpx
import re
import io
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from dotenv import load_dotenv
from bs4 import BeautifulSoup
import calendar

# Load .env from the backend directory
env_path = os.path.join(os.path.dirname(__file__), '.env')
load_dotenv(env_path)

# Try to import supabase, handle gracefully if missing
try:
    from supabase import create_client, Client
    supabase_available = True
except ImportError:
    print("WARNING: supabase module not found. Install with: pip install supabase")
    supabase_available = False
    Client = type('Client', (), {})  # Dummy class
    create_client = None

# API Setup

app = FastAPI(title="ECCHO Social Tracker API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")

supabase: Client = None
local_history = []
if supabase_available and SUPABASE_URL and SUPABASE_KEY and create_client:
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        print("✓ Supabase client initialized successfully")
    except Exception as e:
        print(f"✗ Supabase initialization failed: {e}")
        supabase = None
else:
    print("⚠ Supabase not configured (module missing or env vars not set)")


class SocialLinks(BaseModel):
    brand: str
    facebook_url: Optional[str] = ""
    instagram_url: Optional[str] = ""
    twitter_url: Optional[str] = ""
    linkedin_url: Optional[str] = ""
    youtube_url: Optional[str] = ""


async def fetch_follower_count(url: str, platform: str) -> Optional[int]:
    """Scrape follower count from social media page with LLM fallback."""
    if not url or not url.strip():
        return None

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    }

    def parse_number(text):
        if not text: return None
        
        # 0. Normalize Devanagari digits to standard digits
        devanagari_map = str.maketrans('०१२३४५६७८९', '0123456789')
        text = text.translate(devanagari_map)
        
        # Handle cases like "1.2K", "1,200", "1.2 Lakh", etc.
        # Remove commas and normalize whitespace
        text = text.lower().replace(",", "").strip()
        
        # Remove any non-numeric/period/suffix characters at the end
        # but keep common suffixes (including some Indian language ones)
        # Regex: find a float/int followed by optional suffix
        match = re.search(r'([\d\.]+)\s*([a-z\u0900-\u097F]*)', text)
        if not match: return None
        
        num_str = match.group(1)
        suffix = match.group(2)
        
        multiplier = 1
        # Priority order for suffixes
        if any(x in suffix for x in ['crore', 'cr', 'कोटी', 'करोड']):
            multiplier = 10000000
        elif any(x in suffix for x in ['lakh', 'lac', 'लाख']):
            multiplier = 100000
        elif 'm' in suffix:
            multiplier = 1000000
        elif 'k' in suffix:
            multiplier = 1000
        elif 'b' in suffix:
            multiplier = 1000000000
        elif suffix == 'l': # Single 'l' usually means Lakh in India
            multiplier = 100000
            
        try:
            return int(float(num_str) * multiplier)
        except:
            return None

    # Locale Enforcement for Facebook/Instagram
    if platform in ["facebook", "instagram"]:
        if "?" in url:
            url += "&locale=en_US"
        else:
            url += "?locale=en_US"

    # Platform-specific headers (Bot Emulation)
    if platform in ["facebook", "instagram"]:
        # Use Facebook External Hit User-Agent to bypass login walls
        headers["User-Agent"] = "facebookexternalhit/1.1 (+http://www.facebook.com/externalhit_uatext.php)"
    elif platform == "youtube":
        # Use Googlebot for YouTube
        headers["User-Agent"] = "Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)"

    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            print(f"DEBUG: Scrapping {platform} at {url}")
            resp = await client.get(url, headers=headers)
            
            if resp.status_code != 200: 
                print(f"DEBUG: Scrape failed for {platform} at {url}: Status {resp.status_code}")
                return None
            
            text = resp.text
            print(f"DEBUG: Scraped {len(text)} characters")
            
            # 1. Try Meta Tags (Specifically og:description which is reliable for bots)
            meta_patterns = [
                r'<meta property="og:description" content="([^"]+)"',
                r'<meta name="description" content="([^"]+)"',
                r'<meta name="twitter:description" content="([^"]+)"'
            ]
            for p in meta_patterns:
                m = re.search(p, text, re.I)
                if m:
                    raw_desc = m.group(1)
                    from html import unescape
                    desc = unescape(raw_desc).lower().replace('·', ' ').replace(',', '')
                    print(f"DEBUG: Checking meta: {desc[:80]}...")
                    
                    # Pattern for followers/likes
                    num_match = re.search(r'([\d\.,\u0966-\u096F]+(?:\s*(?:[kmbl]|lakh|crore|lac|लाख|सदस्य))?)\s*(?:followers|subscribers|likes|members|आवडी|फॉलोअर्स|सदस्य)', desc, re.I)
                    if num_match:
                        val = parse_number(num_match.group(1))
                        print(f"DEBUG: Meta match found: {num_match.group(1)} -> {val}")
                        if val: return val

            # 2. Platform Specific JSON/HTML Fallbacks
            if platform == "facebook":
                # Look for direct JSON data
                m = re.search(r'"follower_count":(\d+)', text)
                if m: return int(m.group(1))
                
                patterns = [
                    r'([\d\.,\u0966-\u096F]+[kmb]?)\s*(?:people\s*)?(?:follow|followers|likes|आवडी|फॉलोअर्स|सदस्य)',
                    r'\\"follower_count\\":(\d+)'
                ]
                for p in patterns:
                    m = re.search(p, text, re.I)
                    if m: 
                        val = parse_number(m.group(1))
                        if val: return val
            
            elif platform == "instagram":
                patterns = [
                    r'\\"edge_followed_by\\":\{\\"count\\":(\d+)\}',
                    r'"edge_followed_by":{"count":(\d+)}',
                    r'([\d\.,\u0966-\u096F]+[kmb]?)\s*Followers',
                    r'followed_by\\":(\d+)'
                ]
                for p in patterns:
                    m = re.search(p, text, re.I)
                    if m: 
                        if m.group(1).isdigit(): return int(m.group(1))
                        val = parse_number(m.group(1))
                        if val: return val
            
            elif platform == "youtube":
                patterns = [
                    r'([\d\.,\u0966-\u096F]+[kmb]?)\s*subscribers',
                    r'subscriberCountText.*?label":"([\d\.,\u0966-\u096F]+[kmb]?)\ssubscribers"'
                ]
                for p in patterns:
                    m = re.search(p, text, re.I)
                    if m: 
                        val = parse_number(m.group(1))
                        if val: return val
            
            elif platform == "twitter":
                # Twitter/X is hard to scrape directly. Use syndication endpoint.
                # Extract screen name: https://twitter.com/MDLZ -> MDLZ
                screen_name = url.split('/')[-1].split('?')[0]
                if screen_name:
                    syndication_url = f"https://syndication.twitter.com/srv/timeline-profile/screen-name/{screen_name}"
                    try:
                        s_resp = await client.get(syndication_url, headers=headers)
                        if s_resp.status_code == 200:
                            m = re.search(r'"followers_count":(\d+)', s_resp.text)
                            if m: return int(m.group(1))
                    except: pass
                
                # Fallback to meta tags and text
                patterns = [
                    r'([\d\.,]+[kmb]?)\s*Followers',
                    r'"followers_count":(\d+)',
                    r'<span>([\d\.,]+[kmb]?)\s*Followers</span>'
                ]
                for p in patterns:
                    m = re.search(p, text, re.I)
                    if m:
                        if m.group(1).isdigit(): return int(m.group(1))
                        val = parse_number(m.group(1))
                        if val: return val

            elif platform == "linkedin":
                patterns = [
                    r'([\d\.,\u0966-\u096F]+[kmb]?)\s*followers',
                    r'followerCount":(\d+)',
                    r'content="([\d\.,\u0966-\u096F]+[kmb]?)\s*followers on LinkedIn"'
                ]
                for p in patterns:
                    m = re.search(p, text, re.I)
                    if m:
                        if m.group(1).isdigit(): return int(m.group(1))
                        val = parse_number(m.group(1))
                        if val: return val

    except Exception as e:
        print(f"Error fetching {platform}: {e}")
    return None

def is_last_day_of_month(date_obj):
    last_day = calendar.monthrange(date_obj.year, date_obj.month)[1]
    return date_obj.day == last_day


@app.on_event("startup")
async def startup_event():
    print("\n" + "="*50)
    print("ECCHO SOCIAL MINER API (v2.0 - NEW SCRAPER) IS RUNNING")
    print("="*50 + "\n")

@app.get("/")
async def root():
    return {"status": "ok", "version": "2.0-new-scraper", "message": "ECCHO Social Miner API"}


@app.post("/api/scrape")
async def scrape_followers(links: SocialLinks):
    """Scrape all social media follower counts for a brand."""
    results = {}

    tasks = {
        "facebook": links.facebook_url,
        "instagram": links.instagram_url,
        "twitter": links.twitter_url,
        "linkedin": links.linkedin_url,
        "youtube": links.youtube_url,
    }

    # 1. Validation: ALL provided URLs must be valid for their platforms
    valid_tasks = {}
    domain_map = {
        "facebook": "facebook.com",
        "instagram": "instagram.com",
        "twitter": ["twitter.com", "x.com"],
        "linkedin": "linkedin.com",
        "youtube": ["youtube.com", "youtu.be"]
    }

    for platform, url in tasks.items():
        if url and url.strip():
            allowed = domain_map.get(platform)
            is_valid = False
            if isinstance(allowed, list):
                is_valid = any(d in url.lower() for d in allowed)
            else:
                is_valid = allowed in url.lower()

            if not is_valid:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Invalid URL for {platform.capitalize()}. Please check the link."
                )
            valid_tasks[platform] = url

    if not valid_tasks:
        raise HTTPException(status_code=400, detail="Please provide at least one social media URL.")



    # 3. Run scraping for valid tasks
    for platform, url in valid_tasks.items():
        count = await fetch_follower_count(url, platform)
        results[platform] = count

    # 4. Validation: At least one count must be found
    if not any(v is not None for v in results.values()):
        raise HTTPException(
            status_code=400, 
            detail="Could not extract follower count. The page structure might have changed or the link is private."
        )

    now = datetime.utcnow()
    record = {
        "brand": links.brand,
        "facebook_url": links.facebook_url,
        "instagram_url": links.instagram_url,
        "twitter_url": links.twitter_url,
        "linkedin_url": links.linkedin_url,
        "youtube_url": links.youtube_url,
        "facebook_followers": results.get("facebook"),
        "instagram_followers": results.get("instagram"),
        "twitter_followers": results.get("twitter"),
        "linkedin_followers": results.get("linkedin"),
        "youtube_subscribers": results.get("youtube"),
        "scraped_at": now.isoformat() + "Z",
    }

    if supabase:
        try:
            # 1. Save to raw history
            supabase.table("social_followers").insert(record).execute()
            
            # 2. Logic for reporting tables
            month_name = now.strftime("%B")
            is_last_day = is_last_day_of_month(now)
            is_15th = now.day == 15
            
            report_data = {
                "brand": links.brand,
                "facebook": results.get("facebook"),
                "instagram": results.get("instagram"),
                "twitter": results.get("twitter"),
                "linkedin": results.get("linkedin"),
                "youtube": results.get("youtube"),
            }

            # A. RECURRING TABLE (15th or Last Day)
            # For testing, we allow all days. For production, uncomment the if condition.
            if True: # is_15th or is_last_day: 
                date_label = "15th" if is_15th else (f"{now.day}th" if not is_last_day else "Last Day")
                recurring_record = {
                    **report_data,
                    "month": month_name,
                    "date_label": date_label,
                    "scraped_at": now.isoformat() + "Z"
                }
                supabase.table("recurring_followers").insert(recurring_record).execute()

            # B. AGGREGATE TABLE (Last Day of Month)
            # For testing, we allow all days. For production, uncomment the if condition.
            if True: # is_last_day:
                supabase.table("aggregate_followers").upsert({
                    **report_data,
                    "last_updated": now.date().isoformat()
                }).execute()

        except Exception as e:
            print(f"Database error: {e}")
            raise HTTPException(status_code=500, detail=f"Database update failed: {str(e)}")
    else:
        # Local fallback storage when Supabase is not available
        record["id"] = str(uuid.uuid4())
        local_history.insert(0, record)

    return {"success": True, "brand": links.brand, "data": record}

@app.get("/api/reports/aggregate")
async def get_aggregate_report():
    if not supabase: return {"data": []}
    resp = supabase.table("aggregate_followers").select("*").order("brand").execute()
    return {"data": resp.data}

@app.get("/api/reports/recurring")
async def get_recurring_report():
    if not supabase: return {"data": []}
    resp = supabase.table("recurring_followers").select("*").order("scraped_at", desc=True).execute()
    return {"data": resp.data}


@app.get("/api/dashboard")
async def get_dashboard():
    """Get all saved records from Supabase or local fallback storage."""
    if supabase:
        try:
            resp = supabase.table("social_followers").select("*").order("scraped_at", desc=True).execute()
            return {"data": resp.data}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    return {"data": local_history}


@app.delete("/api/entries/{entry_id}")
async def delete_entry(entry_id: str):
    """Delete a specific record from Supabase or local fallback."""
    if supabase:
        try:
            supabase.table("social_followers").delete().eq("id", entry_id).execute()
            return {"success": True}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    global local_history
    local_history = [r for r in local_history if str(r.get("id")) != str(entry_id)]
    return {"success": True}


@app.delete("/api/entries")
async def delete_all_entries():
    """Clear all records from the feed."""
    if supabase:
        try:
            # Delete all records by filtering on brand (which is always Aditya Birla or others)
            supabase.table("social_followers").delete().neq("brand", "___NONE___").execute()
            return {"success": True}
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    global local_history
    local_history.clear()
    return {"success": True}


@app.get("/api/download-excel")
async def download_excel():
    """Generate and download Excel report from Supabase or local data."""
    records = []
    if supabase:
        try:
            resp = supabase.table("social_followers").select("*").order("scraped_at", desc=True).execute()
            records = resp.data
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    else:
        records = local_history

    wb = openpyxl.Workbook()

    # --- Aggregate Sheet ---
    ws_agg = wb.active
    ws_agg.title = "Aggregate"

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    alt_fill = PatternFill("solid", fgColor="F0F4FF")
    border = Border(
        bottom=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD')
    )

    agg_headers = ["Brand", "Facebook", "Instagram", "X / Twitter", "LinkedIn", "YouTube", "Last Updated"]
    for col, h in enumerate(agg_headers, 1):
        cell = ws_agg.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Group by brand, take latest
    latest = {}
    for r in records:
        b = r.get("brand", "")
        if b not in latest:
            latest[b] = r

    for i, (brand, r) in enumerate(latest.items(), 2):
        row_fill = PatternFill("solid", fgColor="F8F9FF") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [
            brand,
            r.get("facebook_followers") or "",
            r.get("instagram_followers") or "",
            r.get("twitter_followers") or "",
            r.get("linkedin_followers") or "",
            r.get("youtube_subscribers") or "",
            r.get("scraped_at", "")[:10] if r.get("scraped_at") else "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws_agg.cell(row=i, column=col, value=v)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
            cell.border = border
            if isinstance(v, int) and col > 1:
                cell.number_format = '#,##0'

    for col in ws_agg.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws_agg.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
    ws_agg.row_dimensions[1].height = 24

    # --- History Sheet ---
    ws_hist = wb.create_sheet("Full History")
    hist_headers = ["Date", "Brand", "Facebook", "Instagram", "X / Twitter", "LinkedIn", "YouTube"]
    for col, h in enumerate(hist_headers, 1):
        cell = ws_hist.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for i, r in enumerate(records, 2):
        row_fill = PatternFill("solid", fgColor="F8F9FF") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [
            r.get("scraped_at", "")[:10],
            r.get("brand", ""),
            r.get("facebook_followers") or "",
            r.get("instagram_followers") or "",
            r.get("twitter_followers") or "",
            r.get("linkedin_followers") or "",
            r.get("youtube_subscribers") or "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws_hist.cell(row=i, column=col, value=v)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
            cell.border = border
            if isinstance(v, int) and col > 2:
                cell.number_format = '#,##0'

    for col in ws_hist.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws_hist.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"ECCHO_Social_Followers_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )